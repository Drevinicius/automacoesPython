import os
import openpyxl as op
from tkinter import filedialog
from tkinter import Tk, messagebox

class Automacao:
    def caminhoPasta(self):
        root = Tk()
        root.withdraw()

        root.update()
        caminho_planilh_iss = str(filedialog.askopenfilenames(title='Selecione o arquivo do ISS',
                                              initialdir=os.path.expanduser('~')[0:3],
                                             defaultextension='xlsx'))
        root.update()
        caminho_planilh_notas = str(filedialog.askopenfilenames(title='Selecione o arquivo das notas lançadas',
                                              initialdir=os.path.expanduser('~')[0:3],
                                              defaultextension='xlsx'))
        return caminho_planilh_iss, caminho_planilh_notas

    def apuracao_de_iss(self):
        notas_prefeitura, notas_sistema = self.caminhoPasta()

        notas_prefeitura = notas_prefeitura.replace("/", r'\ ')
        notas_prefeitura = notas_prefeitura.replace(' ', r'')
        notas_prefeitura = notas_prefeitura[2:len(notas_prefeitura)-3]

        notas_sistema = notas_sistema.replace('/', r'\ ')
        notas_sistema = notas_sistema.replace(' ', r'')
        notas_sistema = notas_sistema[2:len(notas_sistema)-3]

        try:
            #Abrindo os arquivos
            planilha_notas_prefeitura = op.open(notas_prefeitura)
            planilha_notas_sistema = op.open(notas_sistema)

            #Ativando a planilha na primeira planilha de cada arquivo
            planilha_notas_prefeitura_ativa = planilha_notas_prefeitura[planilha_notas_prefeitura.sheetnames[0]]
            planilha_notas_sistema_ativa = planilha_notas_sistema[planilha_notas_sistema.sheetnames[0]]

            if 'Notas' in planilha_notas_prefeitura.sheetnames:
                ws = planilha_notas_prefeitura['Notas']
                ws.delete_rows(1, 10000)
            else:
                planilha_notas_prefeitura.create_sheet('Notas')
            planilha_relacao = planilha_notas_prefeitura[planilha_notas_prefeitura.sheetnames[1]]

            planilha_relacao.append([])

            total_prefeitura = -1
            total_sistema = 0

            for prefeitura in planilha_notas_prefeitura_ativa.iter_rows(values_only=True):
                nota_existe = False
                for sistema in planilha_notas_sistema_ativa.iter_rows(values_only=True):
                    if prefeitura[1] is None or prefeitura[1] == 'Número' or prefeitura[1] == 'Total de Serviço Tomado':
                        continue
                    if f'{prefeitura[1].strip():0>9}' == sistema[3]:
                        if prefeitura[7] == sistema[16]:
                            nota_existe = True
                            total_sistema += 1

                if prefeitura[6] is not None:
                    total_prefeitura += 1
                if not nota_existe:
                    planilha_relacao.append(prefeitura)

            print(f'Notas da prefeitura: {total_prefeitura}\nNotas no sistema: {total_sistema}')
            planilha_notas_prefeitura.save(notas_prefeitura)

        except PermissionError:
            messagebox.showerror("ARQUIVOS ABERTOS!", "Algum arquivo aberto\nFeche para prosseguir")

if __name__ == '__main__':
    nova = Automacao()
    nova.apuracao_de_iss()


