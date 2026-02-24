import os
import openpyxl as op
from openpyxl.styles import NamedStyle, Font
from tkinter import messagebox
# from datetime import datetime, timedelta

#def somar_dias(data_str=None, dias=7, formato='%d/%m/%Y'):
#   if data_str:
#        data = datetime.strptime(data_str, formato)
#    else:
#        data = datetime.now()

#    nova_data = data + timedelta(days=dias)
#    return nova_data.strftime(formato)

def unirCaminho():
    return os.path.join(os.path.expanduser('~')[0:3], 'relatorio')

def verificacaoPlanilha(pathPlanilha):
    try:
        wb = op.open(pathPlanilha)
        dados = wb['02-03 - Listagem do Browse']

        newSheet = op.Workbook()
        newSheet.create_sheet("dados filtrado")
        newSheet.remove(newSheet['Sheet'])
        planilhaAtiva = newSheet['dados filtrado']

        cabecalho = ['Filial', 'NSU', 'Dt.Emissao', 'Num NF', 'Serie', 'Chave', 'Valor', 'Tipo NF', 'CNPJ/CPF', 'Fornecedor', 'Loja', 'Nome', 'Insc.Estadual', 'Data/Hora Recbto.', 'Situacao NFE', 'Situacao Manifesto', 'Vencimento']
        planilhaAtiva.append(cabecalho)
        interador = 2

        for i in dados.iter_rows(values_only=True, min_row= 4):
           if i[7] == '1 - Saida':
               if i[14] == '1 - Uso Autorizado':
                    if i[15] == '4 - Ciência' or i[15] == '0 - Sem manifestação':
                        planilhaAtiva.append(i)
                        planilhaAtiva[f'C{interador}'].number_format = 'dd/mm/yyyy'
                        planilhaAtiva[f'G{interador}'].number_format = 'R$ #,##0.00_);[Red](R$ #,##0.00)'
                        planilhaAtiva[f'Q{interador}'].number_format = 'dd/mm/yyyy'
                        planilhaAtiva[f'Q{interador}'] = f'= C{interador} + 45'

                        interador += 1

        pathNovo = os.path.join(unirCaminho(), 'dadosXML.xlsx')
        newSheet.save(pathNovo)
    except PermissionError:
        messagebox.showerror("Arquivo aberto", "Feche a planilha 'dadosXML.xlsl' e tente novamente")
    else:
        messagebox.showinfo("Processo finalizado", "Arquivo gerado com sucesso")

path = unirCaminho()

if not os.path.exists(path):
    os.mkdir(path)
else:
    path = os.path.join(path, "smxmlcentrl.xlsx")

if os.path.exists(path):
    messagebox.showinfo("Arquivo encontrato","Click em 'ok' para iniciar o processo")
    verificacaoPlanilha(path)
else:
    messagebox.showerror("Arquivo não encontrato","renomei o arquivo como 'smxmlcentrl' ou adicione a pasta relatorio")

