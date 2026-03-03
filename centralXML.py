import os
import openpyxl
from openpyxl.styles import NamedStyle, Font
from tkinter import messagebox
from datetime import datetime, timedelta


class Automacao:
    def comparar_data(self, data_verificar, formato='%d/%m/%Y'):
        # Converter data de verificação
        data_verificar = datetime.strptime(data_verificar, formato)
        # Calcular data atual + 7 dias
        data_limite = datetime.now() + timedelta(days=7)
        # Comparar
        return data_verificar < data_limite

    def caminhosDoArquivo(self):
        pathArquivo = pathSalvarArquivo = os.path.join(os.path.expanduser('~')[0:3], 'relatorio')

        pathSalvarArquivo = os.path.join(pathSalvarArquivo,'dadosXML.xlsx')
        pathArquivo = os.path.join(pathArquivo, "smxmlcentrl.xlsx")

        if not os.path.exists(os.path.join(os.path.expanduser('~')[0:3], 'relatorio')):
            os.makedirs(os.path.join(os.path.expanduser('~')[0:3], 'relatorio'))
            pathArquivo = None

        return pathArquivo, pathSalvarArquivo

    def manipular_sheet(self):
        try:
            pathPlanilha, pathSalvarPlanilha = self.caminhosDoArquivo()

            # Abrir a planilha de onde serão extraídos os meus dados
            wb = openpyxl.open(pathPlanilha)
            dados_extraidos = wb[wb.sheetnames[0]]

            # Criando a nova planilha que vão receber os meus dados
            novaPlanilha = openpyxl.Workbook()
            novaPlanilha.create_sheet('dados filtrados')
            novaPlanilha.remove(novaPlanilha['Sheet'])

            planilha = novaPlanilha['dados filtrados']
            cabecalho = ['Filial', 'NSU', 'Dt.Emissao', 'Num NF', 'Serie', 'Chave', 'Valor', 'Tipo NF', 'CNPJ/CPF',
                         'Fornecedor', 'Loja', 'Nome', 'Insc.Estadual', 'Data/Hora Recbto.', 'Situacao NFE',
                         'Situacao Manifesto', 'Vencimento']
            planilha.append(cabecalho)
            interador_de_adicao = 2

            for i in dados_extraidos.iter_rows(values_only=True):
                if i[7] == '1 - Saida':
                    if i[14] == '1 - Uso Autorizado':
                        if i[15] == '4 - Ciência' or i[15] == '0 - Sem manifestação':
                            planilha.append(i)
                            planilha[f'C{interador_de_adicao}'].number_format = 'dd/mm/yyyy'
                            planilha[f'G{interador_de_adicao}'].number_format = 'R$ #,##0.00_);[Red](R$ #,##0.00)'
                            planilha[f'Q{interador_de_adicao}'] = planilha[f'C{interador_de_adicao}'].value + timedelta(days=45)
                            planilha[f'Q{interador_de_adicao}'].number_format = 'dd/mm/yyyy'

                            if self.comparar_data(planilha[f'Q{interador_de_adicao}'].value.strftime('%d/%m/%Y')):
                                planilha[f'Q{interador_de_adicao}'].font = Font(bold=True)

                            interador_de_adicao += 1

            novaPlanilha.save(pathSalvarPlanilha)
        except (FileNotFoundError, PermissionError):
            messagebox.showerror("ARQUIVO NÃO ENCONTRADO/ABERTO", f"""VOCÊ DEVE RENOMEAR O ARQUIVO COMO 'smxmlcentrl.xlsx,
SALVE NO DIRETÓRIO {os.path.expanduser('~')[:3]}relatorio,
OU FECHE 'dadosXML.xlsx'""")


